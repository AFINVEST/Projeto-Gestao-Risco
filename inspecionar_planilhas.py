"""
inspecionar_planilhas.py
=========================
Inspeciona cada planilha Excel usada no fluxo e dumpa:
  - Lista de sheets
  - Colunas e dimensoes de cada sheet
  - Primeiras/ultimas 5 linhas
  - Tipo de valores (numerico, string, formula-refresh?)

Nao modifica NADA. So le e imprime.
"""
from pathlib import Path
import pandas as pd
import openpyxl

PLANILHAS = [
    ("Dados/BBG - ECO DASH.xlsx",                                        ["BZ RATES", "DIV01"]),
    ("Dados/AF_Trading.xlsm",                                            ["Base CDI", "Base IPCA"]),
    ("Dados/FechamentoNTNBs.xlsx",                                       None),
    ("Dados/DadosJuros.xlsx",                                            None),
    (r"Z:\Asset Management\FUNDOS e CLUBES\Gerencial\dashboard LFT.xlsx", ["Historico preços"]),
]


def _tem_formulas(path, sheet):
    """Detecta se a sheet tem formulas Bloomberg (=BDH, =BDP, etc)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        formulas_bbg = 0
        outras_formulas = 0
        for row in ws.iter_rows(max_row=50, values_only=False):
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                v = cell.value.strip()
                if v.startswith("="):
                    if any(f in v.upper() for f in ("BDH", "BDP", "BDS", "BLPAPI")):
                        formulas_bbg += 1
                    else:
                        outras_formulas += 1
        return {"bbg": formulas_bbg, "outras": outras_formulas}
    except Exception as e:
        return f"erro: {e}"


def inspecionar(path_str, sheets_alvo=None):
    path = Path(path_str)
    print("\n" + "=" * 90)
    print(f"ARQUIVO: {path_str}")
    print("=" * 90)

    if not path.exists():
        print(f"[erro] arquivo nao encontrado")
        return

    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        print(f"[erro] nao consegui abrir: {e}")
        return

    print(f"Sheets disponiveis: {xl.sheet_names}")
    print(f"Tamanho do arquivo: {path.stat().st_size / 1024:.1f} KB")
    print(f"Ultima modificacao: {pd.Timestamp(path.stat().st_mtime, unit='s')}")

    alvo = sheets_alvo if sheets_alvo else xl.sheet_names

    for sh in alvo:
        if sh not in xl.sheet_names:
            print(f"\n  [SKIP] sheet '{sh}' nao existe")
            continue
        try:
            df = xl.parse(sh)
        except Exception as e:
            print(f"\n  [erro parse sheet '{sh}']: {e}")
            continue

        print(f"\n--- Sheet: '{sh}' " + "-" * 40)
        print(f"  Dimensoes: {df.shape[0]} linhas x {df.shape[1]} colunas")
        print(f"  Colunas: {list(df.columns)[:10]}{'... (+' + str(len(df.columns)-10) + ')' if len(df.columns) > 10 else ''}")

        # Formulas
        f = _tem_formulas(path, sh)
        if isinstance(f, dict):
            print(f"  Formulas Bloomberg (BDH/BDP): {f['bbg']}  | Outras formulas: {f['outras']}")

        # Amostra
        print(f"\n  Primeiras 3 linhas:")
        print(df.head(3).to_string(max_cols=8, max_colwidth=25))
        print(f"\n  Ultimas 3 linhas:")
        print(df.tail(3).to_string(max_cols=8, max_colwidth=25))


if __name__ == "__main__":
    for path, sheets in PLANILHAS:
        inspecionar(path, sheets)
    print("\n" + "=" * 90)
    print("Fim da inspecao.")
